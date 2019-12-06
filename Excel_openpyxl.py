# encoding: utf-8
"""
Author: 沙振宇
CreateTime: 2019-12-6
UpdateTime: 2019-12-6
Info: 可读写Excel2007 XLSX/XLSM文件, 不可以处理XLS文件
"""
import openpyxl

def write_excel():
    f = openpyxl.Workbook()  # 创建工作簿
    # sheet1 = f.create_sheet()
    print("sheet names:", f.sheetnames)
    sheet1 = f.active   # 当前工作表的名称
    print("sheet1:", sheet1)
    row_file = 5 # 生成5行
    col_three = 3 # 生成3列

    for row in range(row_file):
        for col in range(col_three):
            rw = row + 1
            cl = col + 1
            if col % 2 == 0:
                sheet1.cell(row=rw, column=cl, value='1')
            else:
                sheet1.cell(row=rw, column=cl, value='2')

    f.save("file/生成的Excel_OPENPYXL.xlsx")  # 保存文件

if __name__ == '__main__':
    # 写入Excel
    write_excel()
    print('openpyxl 写入成功')